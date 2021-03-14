# -*- coding: UTF-8 -*-
import xlrd
import xlwt
import sys
import logging
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
import tkinter.messagebox
############ 配置 ################

############ 配置 ################

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("excel_test")


class ExcelReadUtil:
    def __init__(self, filename, sheet_index=0, is_has_title=True):
        """
        excel文件对象，类克返回每一列数据，可选择类型：(int float string)
        :param filename:  excel 文件名
        :param sheet_index: 文件中的第几个sheet 下标从零开始
        :param is_has_title: 是否有标题行
        """
        input_excel = xlrd.open_workbook(filename).sheets()
        assert len(input_excel) > sheet_index, "sheet 编号超出sheet数量！"
        sheet = input_excel[sheet_index]
        row_count = sheet.nrows
        col_count = sheet.ncols
        # 几列就是几个列表
        # self.data = [[""] * row_count] * col_count
        self.data = [["" for i in range(row_count)] for i in range(col_count)]

        self.title = [None] * col_count
        for col in range(col_count):
            for i in range(1 if is_has_title else 0, row_count):
                self.data[col][i-1] = str(sheet.cell(i, col).value)
        if is_has_title:
            for col in range(col_count):
                self.title[col] = str(sheet.cell(0, col).value)

    def get_col_str(self, col_num):
        return self.data[col_num]

    def get_col_int(self, col_num):
        try:
            return [int(item) for item in self.data[col_num]]
        except Exception as e:
            logger.error("转换int 类型失败！ e:{}".format(col_num))

    def get_col_float(self, col_num):
        try:
            return [float(item) for item in self.data[col_num]]
        except Exception as e:
            logger.error("转换int 类型失败！ e:{}".format(e))


class ExcelWriteUtil(object):
    @classmethod
    def write_excel(cls, filename, data, sheet_name="sheet0", title=None):
        """
        写excel文件
        :param filename:  文件名，可以是全路径
        :param title: 标题行（第一行）,没有填 None
        :param data: 数据，二维列表
        """
        save_excel = xlwt.Workbook()
        out_sheet = save_excel.add_sheet(sheet_name, cell_overwrite_ok=True)  # 创建sheet
        col_num = len(data)
        row_num = len(data[0])
        # 保证列数相同
        if title is not None:
            assert len(title) == len(data)
        for i in range(col_num):
            if title is None:
                break
            out_sheet.write(0, i, title[i])
        for row in range(row_num):
            for col in range(col_num):
                out_sheet.write(row, col, data[col][row])
        save_excel.save(filename)

    @classmethod
    def write_excel2(cls, filename, data, sheet_name="sheet0", title=None):
        """
        写excel文件
        :param filename:  文件名，可以是全路径
        :param title: 标题行（第一行）,没有填 None
        :param data: 数据，二维列表
        """
        wb = Workbook()
        ws = wb.active
        wb.active.title = sheet_name
        col_num = len(data)
        row_num = len(data[0])
        if title is not None:
            assert len(title) == len(data)
            ws.append(title)
        for row in range(row_num):
            temp_row = []
            for col in range(col_num):
                temp_row.append(data[col][row])
            ws.append(temp_row)
        wb.save(filename)

# if __name__ == "__main__":
#      filename = "a.xlsx"
#      excel_list = ExcelReadUtil(filename)
#      logger.info(excel_list.title)
#      # logger.info(excel_list.data[3])
#      ExcelWriteUtil.write_excel2("save.xlsx", excel_list.data,"sheet",excel_list.title)
